# preparation/views.py
#from logging.config import stopListening
import openpyxl
#from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
#from django.template.defaulttags import comment
from django.core.exceptions import ObjectDoesNotExist

from core.models import *
#from django.contrib.auth.models import Group
from django.db import connection, connections, transaction
from django.contrib import messages
#from django.http import JsonResponse

from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
import logging
logger = logging.getLogger(__name__)


def preparation_dashboard(request):
    # is_preparation = request.user.groups.filter(name='preparations').exists()
    # return render(request, 'preparation/preparation_dashboard.html', {'is_preparation': is_preparation})
    return render(request, 'preparation/preparation_dashboard.html')

def po_stock(request):
    # return HttpResponse("po_stock view is working!")
    with connections['default'].cursor() as cursor:
        cursor.execute("""SELECT  
    pos.id,
    pos.po,
    posum.location_all,
    pos.size,
    pos.style,
    pos.color,
    pos.color_desc,
    pos.flash,
    pos.brand,
    pos.skeda,
    pos.total_order_qty,
    pos.no_lines_by_skeda,
    pos.hangtag,
    p.location,

    bs.stock_b,
    br.request_b,
    cs.stock_c,
    cr.request_c

FROM pos

LEFT JOIN prep_locations AS p ON p.id = pos.loc_id_su
LEFT JOIN [172.27.161.200].[posummary].dbo.pro AS posum ON posum.po_new = pos.po_new

-- Pre-aggregated subqueries
LEFT JOIN (
    SELECT po_id, SUM(qty) AS stock_b
    FROM barcode_stocks
    GROUP BY po_id
) AS bs ON bs.po_id = pos.id

LEFT JOIN (
    SELECT po_id, SUM(qty) AS request_b
    FROM barcode_requests
    WHERE status != 'error'
    GROUP BY po_id
) AS br ON br.po_id = pos.id

LEFT JOIN (
    SELECT po_id, SUM(qty) AS stock_c
    FROM carelabel_stocks
    GROUP BY po_id
) AS cs ON cs.po_id = pos.id

LEFT JOIN (
    SELECT po_id, SUM(qty) AS request_c
    FROM carelabel_requests
    WHERE status != 'error'
    GROUP BY po_id
) AS cr ON cr.po_id = pos.id

WHERE pos.closed_po = 'Open'

ORDER BY pos.po ASC, pos.size DESC""")
        lines = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in lines]

        for row in data:
            total_order_qty = row.get('total_order_qty') or 0
            row['95'] = round(total_order_qty * 0.95,1)

            stock_b = row.get('stock_b') or 0
            request_b = row.get('request_b') or 0
            row['stock_percentage_b'] = round((stock_b / total_order_qty * 100), 2) if total_order_qty else 0
            row['to_print_b'] = total_order_qty - stock_b
            row['on_stock_b'] = stock_b - request_b
            row['request_b'] = request_b

            stock_c = row.get('stock_c') or 0
            request_c = row.get('request_c') or 0
            row['stock_percentage_c'] = round((stock_c / total_order_qty * 100), 2) if total_order_qty else 0
            row['to_print_c'] = total_order_qty - stock_c
            row['on_stock_c'] = stock_c - request_c
            row['request_c'] = request_c

    return render(request, 'preparation/po_stock.html', {'data': data})

def barcode_requests(request, id=None, action=None):
    success_msg = ""
    errors = []

    if request.method == 'POST':
        # Handle form submission
        id = request.POST.get('id')
        qty = request.POST.get('qty')
        comment = request.POST.get('comment')

        try:
            request_data = BarcodeRequests.objects.get(id=id)
            request_data.qty = int(qty)
            request_data.comment = comment
            request_data.status = 'confirmed'
            request_data.save()

            success_msg = "Request confirmed successfully."
        except BarcodeRequests.DoesNotExist:
            errors.append("Request not found.")
        except ValueError:
            errors.append("Invalid quantity.")

        #stay on same page
        # return render(request, 'preparation/barcode_requests.html', {
        #     'id': id,
        #     'request_data': request_data,
        #     'success_msg': success_msg,
        #     'errors': errors
        # })

    elif id is not None:
        request_data = get_object_or_404(BarcodeRequests, id=id)

        if action == "edit":
            return render(request, 'preparation/barcode_requests.html', {
                'id': id,
                'request_data': request_data
            })

        elif action == "error":
            # return HttpResponse("✅ Request marked as error.")

            request_data.status = 'error'
            request_data.qty = 0
            request_data.save()

            success_msg = "Request canceled"

        elif action == "rfid":
            # return HttpResponse("✅ RFID request completed.")

            request_data.status = 'done'
            request_data.qty = 0
            request_data.save()

            success_msg = "RFID request confirmed successfully."

        elif action == "print":
            # return HttpResponse("🖨️ Print action executed.")
            with connections['default'].cursor() as cursor:
                cursor.execute("""SELECT r.po_id,r.user_id,r.ponum,r.size,r.qty,r.module,
                        r.leader,r.status,r.type,r.comment,r.created_at,p.style,p.color
                    FROM barcode_requests AS r
                    JOIN pos AS p ON p.po = r.ponum
                    WHERE r.id = %s
                """, [id])

                row = cursor.fetchone()
                if not row:
                    return HttpResponse("❌ Request not found.")
                columns = [col[0] for col in cursor.description]
                request_new = dict(zip(columns, row))

            PrintRequestLabels.objects.create(
                po_id=request_new['po_id'],
                po=request_new['ponum'],
                type='Barcode',
                style=request_new.get('style'),
                color=request_new.get('color'),
                size=request_new.get('size'),
                module=request_new.get('module'),
                leader=request_new.get('leader'),
                comment=request_new.get('comment'),
                created=request_new['created_at'].strftime('%Y-%m-%d %H:%M') if request_new['created_at'] else '',
                printer="Preparacija Zebra",
                qty=request_new['qty']
            )

            success_msg = "Request sent to printer."

    # Show table of requests
    with connections['default'].cursor() as cursor:
        cursor.execute("""
            SELECT 
    r.id,
    r.po_id,
    r.user_id,
    r.ponum,
    r.size,
    r.module,
    r.leader,
    r.status,
    r.type,
    r.comment,
    r.qty,
    r.created_at,
    r.updated_at,
    pos.total_order_qty,
    pos.style,
    pos.color,
    pos.po_new,
    ISNULL(bs.stock_qty, 0) AS stocks,
    ISNULL(br.request_qty, 0) AS requests
FROM barcode_requests AS r
JOIN pos ON pos.id = r.po_id

-- Pre-aggregated stock sums
LEFT JOIN (
    SELECT po_id, SUM(qty) AS stock_qty
    FROM barcode_stocks
    GROUP BY po_id
) AS bs ON bs.po_id = r.po_id

-- Pre-aggregated request sums (excluding 'error')
LEFT JOIN (
    SELECT po_id, SUM(qty) AS request_qty
    FROM barcode_requests
    WHERE status != 'error'
    GROUP BY po_id
) AS br ON br.po_id = r.po_id

WHERE 
    (CAST(r.created_at AS DATE) = CAST(GETDATE() AS DATE) AND r.status NOT IN ('error'))
    OR r.status = 'pending'
ORDER BY 
    r.status DESC,
    r.created_at ASC;
        """)

        lines = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in lines]

    for row in data:
        total_order_qty = row.get('total_order_qty') or 0
        stock_b = row.get('stocks') or 0
        request_b = row.get('requests') or 0
        row['to_print_b'] = total_order_qty - stock_b
        row['on_stock_b'] = stock_b - request_b

    return render(request, 'preparation/barcode_requests.html', {
        'data': data,
        'success_msg': success_msg,
        'errors': errors
    })

def carelabel_requests(request, id=None, action=None):
    success_msg = ""
    errors = []

    if request.method == 'POST':
        # Handle form submission
        id = request.POST.get('id')
        qty = request.POST.get('qty')
        comment = request.POST.get('comment')

        try:
            request_data = CarelabelRequests.objects.get(id=id)
            request_data.qty = int(qty)
            request_data.comment = comment
            request_data.status = 'confirmed'
            request_data.save()

            success_msg = "Request confirmed successfully."
        except CarelabelRequests.DoesNotExist:
            errors.append("Request not found.")
        except ValueError:
            errors.append("Invalid quantity.")

        #stay on same page
        # return render(request, 'preparation/carelabel_requests.html', {
        #     'id': id,
        #     'request_data': request_data,
        #     'success_msg': success_msg,
        #     'errors': errors
        # })

    elif id is not None:
        request_data = get_object_or_404(CarelabelRequests, id=id)

        if action == "edit":
            return render(request, 'preparation/carelabel_requests.html', {
                'id': id,
                'request_data': request_data
            })

        elif action == "error":
            # return HttpResponse("✅ Request marked as error.")

            request_data.status = 'error'
            request_data.qty = 0
            request_data.save()

            success_msg = "Request canceled"

        elif action == "rfid":
            # return HttpResponse("✅ RFID request completed.")

            request_data.status = 'done'
            request_data.qty = 0
            request_data.save()

            success_msg = "RFID request confirmed successfully."

        elif action == "print":
            # return HttpResponse("🖨️ Print action executed.")
            with connections['default'].cursor() as cursor:
                cursor.execute("""SELECT r.po_id,r.user_id,r.ponum,r.size,r.qty,r.module,
                        r.leader,r.status,r.type,r.comment,r.created_at,p.style,p.color
                    FROM carelabel_requests AS r
                    JOIN pos AS p ON p.po = r.ponum
                    WHERE r.id = %s
                """, [id])

                row = cursor.fetchone()
                if not row:
                    return HttpResponse("❌ Request not found.")
                columns = [col[0] for col in cursor.description]
                request_new = dict(zip(columns, row))

            PrintRequestLabels.objects.create(
                po_id=request_new['po_id'],
                po=request_new['ponum'],
                type='Carelabel',
                style=request_new.get('style'),
                color=request_new.get('color'),
                size=request_new.get('size'),
                module=request_new.get('module'),
                leader=request_new.get('leader'),
                comment=request_new.get('comment'),
                created=request_new['created_at'].strftime('%Y-%m-%d %H:%M') if request_new['created_at'] else '',
                printer="Preparacija Zebra",
                qty=request_new['qty']
            )

            success_msg = "Request sent to printer."

    # Show table of requests
    with connections['default'].cursor() as cursor:
        cursor.execute("""
            SELECT 
    r.id,
    r.po_id,
    r.user_id,
    r.ponum,
    r.size,
    r.module,
    r.leader,
    r.status,
    r.type,
    r.comment,
    r.qty,
    r.created_at,
    r.updated_at,
    pos.total_order_qty,
    pos.style,
    pos.color,
    pos.po_new,
    ISNULL(cs.stock_qty, 0) AS stocks,
    ISNULL(cr.request_qty, 0) AS requests
FROM carelabel_requests AS r
JOIN pos ON pos.id = r.po_id

-- Pre-aggregated stock sums
LEFT JOIN (
    SELECT po_id, SUM(qty) AS stock_qty
    FROM carelabel_stocks
    GROUP BY po_id
) AS cs ON cs.po_id = r.po_id

-- Pre-aggregated request sums (excluding 'error')
LEFT JOIN (
    SELECT po_id, SUM(qty) AS request_qty
    FROM carelabel_requests
    WHERE status != 'error'
    GROUP BY po_id
) AS cr ON cr.po_id = r.po_id

WHERE 
    (CAST(r.created_at AS DATE) = CAST(GETDATE() AS DATE) AND r.status NOT IN ('error'))
    OR r.status = 'pending'
ORDER BY 
    r.status DESC,
    r.created_at ASC;
""")

        lines = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in lines]

    for row in data:
        total_order_qty = row.get('total_order_qty') or 0
        stock_c = row.get('stocks') or 0
        request_c = row.get('requests') or 0
        row['to_print_c'] = total_order_qty - stock_c
        row['on_stock_c'] = stock_c - request_c

    return render(request, 'preparation/carelabel_requests.html', {
        'data': data,
        'success_msg': success_msg,
        'errors': errors
    })

def secondq_requests(request, id=None, action=None):
    success_msg = ""
    errors = []

    if request.method == 'POST':
        # Handle form submission
        id = request.POST.get('id')
        qty = request.POST.get('qty')
        comment = request.POST.get('comment')

        try:
            request_data = SecondQRequests.objects.get(id=id)
            request_data.qty = int(qty)
            request_data.comment = comment
            request_data.status = 'confirmed'
            request_data.save()

            success_msg = "Request confirmed successfully."
        except CarelabelRequests.DoesNotExist:
            errors.append("Request not found.")
        except ValueError:
            errors.append("Invalid quantity.")

        #stay on same page
        # return render(request, 'preparation/secondq_requests.html', {
        #     'id': id,
        #     'request_data': request_data,
        #     'success_msg': success_msg,
        #     'errors': errors
        # })

    elif id is not None:

        if action == "edit":
            request_data = get_object_or_404(SecondQRequests, id=id)
            return render(request, 'preparation/secondq_requests.html', {
                'id': id,
                'request_data': request_data
            })

        elif action == "error":
            # return HttpResponse("✅ Request marked as error.")
            request_data = get_object_or_404(SecondQRequests, id=id)

            request_data.status = 'error'
            request_data.qty = 0
            request_data.save()

            success_msg = "Request canceled"

        elif action == "rfid":
            # return HttpResponse("✅ RFID request completed.")
            request_data = get_object_or_404(SecondQRequests, id=id)

            request_data.status = 'done'
            request_data.qty = 0
            request_data.save()

            success_msg = "Second quality request confirmed successfully."

        elif action == "print":
            request_data = get_object_or_404(SecondQRequests, id=id)
            # return HttpResponse("🖨️ Print action executed.")
            with connections['default'].cursor() as cursor:
                cursor.execute("""SELECT r.po_id,r.user_id,r.ponum,r.size,r.qty,r.module,
                        r.leader,r.status,r.type,r.comment,r.created_at,p.style,p.color
                    FROM secondq_requests AS r
                    JOIN pos AS p ON p.po = r.ponum
                    WHERE r.id = %s
                """, [id])

                row = cursor.fetchone()
                if not row:
                    return HttpResponse("❌ Request not found.")
                columns = [col[0] for col in cursor.description]
                request_new = dict(zip(columns, row))


        elif action == "update":
            return HttpResponse("Update action executed.")

            pending_requests = SecondQRequests.objects.using('default').filter(status='pending')

            for row in pending_requests:
                style = row.style
                color = row.color
                size = row.size

                with connections['default'].cursor() as cursor:
                    cursor.execute("""
                                SELECT [Item No_], [Color], [TG],
                                       [Materiale] as materiale,
                                       [Description Model] as des,
                                       [TG2] as tg2,
                                       [Commersial Color code] as ccc,
                                       [Color decstionption] as cd,
                                       [Barcode] as barcode
                                FROM [172.27.161.200].[preparation].[dbo].[Barcode Table Quality]
                                WHERE [Item No_] = %s AND [Color] = %s AND [TG] = %s
                            """, [style, color, size])

                    po = cursor.fetchone()

                if not po:
                    msg = f"Problem to find in table: {style} {color} {size}, contact Italy and call Aslan :)"
                    return HttpResponse(f"<h3 style='color:red'>{msg}</h3>")

                materiale, desc, tg2, ccc, cd, barcode = po[3], po[4], po[5], po[6], po[7], po[8]

                # Step 2: Update the record
                row.materiale = materiale
                row.desc = desc
                row.tg2 = tg2
                row.ccc = ccc
                row.cd = cd
                row.barcode = barcode
                row.save(using='default')

            success_msg = "Second quality requests are updated."

    # Show table of requests
    with connections['default'].cursor() as cursor:
        cursor.execute("""
            SELECT 
    r.id,
    r.po_id,
    r.user_id,
    r.ponum,
    r.size,
    r.module,
    r.leader,
    r.status,
    r.type,
    r.comment,
    r.qty,
    r.created_at,
    r.updated_at,
    r.style,
    r.color,
    r.materiale,
    r.tg2,
    r.[desc],
    r.ccc,
    r.cd,
    r.barcode,
    pos.total_order_qty,
    pos.po_new,
    ISNULL(sr.stock_qty, 0) AS stocks,
    ISNULL(sr.request_qty, 0) AS requests
FROM secondq_requests AS r
JOIN pos ON pos.id = r.po_id

-- Pre-aggregated stock and request sums
LEFT JOIN (
    SELECT 
        po_id,
        SUM(qty) AS stock_qty,
        SUM(CASE WHEN status != 'error' THEN qty ELSE 0 END) AS request_qty
    FROM secondq_requests
    GROUP BY po_id
) AS sr ON sr.po_id = r.po_id

WHERE 
    (CAST(r.created_at AS DATE) = CAST(GETDATE() AS DATE) AND r.status NOT IN ('error'))
    OR r.status = 'pending'
ORDER BY 
    r.status DESC,
    r.created_at ASC
        """)

        lines = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in lines]

    for row in data:
        total_order_qty = row.get('total_order_qty') or 0
        stock_c = row.get('stocks') or 0
        request_c = row.get('requests') or 0
        row['to_print_c'] = total_order_qty - stock_c
        row['on_stock_c'] = stock_c - request_c

    # return HttpResponse(f"{data}")

    return render(request, 'preparation/secondq_requests.html', {
        'data': data,
        'success_msg': success_msg,
        'errors': errors
    })

def functions(request):
    # return HttpResponse("Stock view is working!")

    return render(request, 'preparation/functions.html')

def add_to_stock(request):
    # return HttpResponse("add_to_stock view is working!")
    errors = []
    success_msg = ""

    if request.method == 'POST':
        # print(request.POST)
        po_num = request.POST.get('po')
        qty = int(request.POST.get('qty'))
        barcode = request.POST.get('barcode', '0')
        carelabel = request.POST.get('carelabel', '0')
        machine = request.POST.get('machine')
        machine_c = request.POST.get('machine_c')
        comment = request.POST.get('comment', '')

        # Validate required fields
        if not po_num or len(po_num) < 6 or len(po_num) > 7:
            errors.append("PO number must be 6-7 characters long")

        if not qty:
            errors.append("Qty/Kolicina je obavezna")

        # Verify if PO exists and is not closed
        try:
            po = Pos.objects.get(po=po_num)
        except ObjectDoesNotExist:
            errors.append("Komesa doesn't exist in the PO table")
        else:
            if po.closed_po == 'Closed':
                errors.append("PO is Closed")

        # Handle barcode
        if barcode != '0':
            if not machine:
                errors.append("Barcode masina nije izabrana")

            if not errors:  # Proceed if no errors so far
                try:
                    BarcodeStocks.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=qty,
                        type="new",
                        comment=comment,
                        machine=machine
                    )
                    success_msg += "BarcodeStocks uspesno snimljen. <br>"  # Append the success message
                except Exception as e:
                    errors.append("Problem saving to BarcodeStocks table")

        # Handle carelabel
        if carelabel != '0':
            if not machine_c:
                errors.append("Carelabel type nije izabran")

            if not errors:  # Proceed if no errors so far
                try:
                    CarelabelStocks.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=qty,
                        type="new",
                        comment=comment,
                        machine=machine_c
                    )
                    success_msg += "CarelabelStocks uspesno snimljen."  # Append the success message
                except Exception as e:
                    errors.append("Problem saving to CarelabelStocks table")

        if carelabel == '0' and barcode == '0':
            errors.append("Nije oznacen ni barcode ni carelabel")

        # If there are no errors, pass success_msg to template
        if not errors:
            return render(request, 'preparation/add_to_stock.html', {
                'pos': Pos.objects.filter(closed_po='Open'),
                'success_msg': success_msg
            })

        # If errors exist, pass them to the template
        return render(request, 'preparation/add_to_stock.html', {
            'pos': Pos.objects.filter(closed_po='Open'),
            'errors': errors
        })

    # If GET request, just render the form with open POS
    return render(request, 'preparation/add_to_stock.html', {'pos': Pos.objects.filter(closed_po='Open')})

def back_from_module(request):
    # return HttpResponse("back_from_module view is working!")

    # find list of modules/lines
    with connections['bbstock_db'].cursor() as cursor:
        cursor.execute("SELECT [location] as line FROM [bbStock].[dbo].[locations]")
        lines = cursor.fetchall()
    lines = [row[0] for row in lines]
    # print(lines)

    errors = []
    success_msg = ""

    if request.method == 'POST':
        # print(request.POST)
        po_num = request.POST.get('po')
        qty = request.POST.get('qty')
        barcode = request.POST.get('barcode', '0')
        carelabel = request.POST.get('carelabel', '0')
        modul = request.POST.get('modul')
        comment = request.POST.get('comment', '')

        # Validate required fields
        if not po_num or len(po_num) < 6 or len(po_num) > 7:
            errors.append("PO number must be 6-7 characters long")

        if not qty:
            errors.append("Qty/Kolicina je obavezna")

        if not modul:
            errors.append("Modul/line polje je obavezno")

        # Verify if PO exists and is not closed
        try:
            po = Pos.objects.get(po=po_num)
        except ObjectDoesNotExist:
            errors.append("Komesa doesn't exist in the PO table")
        else:
            if po.closed_po == 'Closed':
                errors.append("PO is Closed")

        # Handle barcode
        if barcode != '0':

            if not errors:  # Proceed if no errors so far
                try:
                    BarcodeRequests.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty) * (-1),  # negative qty
                        module=modul,
                        type="modul",
                        status="back",
                        comment=comment,

                    )
                    success_msg += "BarcodeRequests uspesno snimljen. <br>"  # Append the success message
                except Exception as e:
                    errors.append("Problem saving to BarcodeRequests table")

        # Handle carelabel
        if carelabel != '0':

            if not errors:  # Proceed if no errors so far
                try:
                    CarelabelRequests.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty) * (-1),  # negative qty
                        module=modul,
                        type="modul",
                        status="back",
                        comment=comment,

                    )
                    success_msg += "CarelabelRequests uspesno snimljen."  # Append the success message
                except Exception as e:
                    errors.append("Problem saving to CarelabelRequests table")

        if carelabel == '0' and barcode == '0':
            errors.append("Nije oznacen ni barcode ni carelabel")

        # If there are no errors, pass success_msg to template
        if not errors:
            return render(request, 'preparation/back_from_module.html', {
                'pos': Pos.objects.filter(closed_po='Open'),
                'lines': lines,
                'success_msg': success_msg
            })

        # If errors exist, pass them to the template
        return render(request, 'preparation/back_from_module.html', {
            'pos': Pos.objects.filter(closed_po='Open'),
            'lines': lines,
            'errors': errors
        })


    # If GET request, just render the form with open POS
    return render(request, 'preparation/back_from_module.html', {
        'pos': Pos.objects.filter(closed_po='Open'),
        'lines': lines
    })

def reduce_from_stock(request):
    # return HttpResponse("reduce_from_stock view is working!")

    errors = []
    success_msg = ""

    if request.method == 'POST':
        # print(request.POST)
        po_num = request.POST.get('po')
        qty = request.POST.get('qty')
        barcode = request.POST.get('barcode', '0')
        carelabel = request.POST.get('carelabel', '0')
        comment = request.POST.get('comment', '')

        # Validate required fields
        if not po_num or len(po_num) < 6 or len(po_num) > 7:
            errors.append("PO number must be 6-7 characters long")

        if not qty:
            errors.append("Qty/Kolicina je obavezna")

        # Verify if PO exists and is not closed
        try:
            po = Pos.objects.get(po=po_num)
        except ObjectDoesNotExist:
            errors.append("Komesa doesn't exist in the PO table")
        else:
            if po.closed_po == 'Closed':
                errors.append("PO is Closed")

        # Handle barcode
        if barcode != '0':

            if not errors:  # Proceed if no errors so far
                try:
                    BarcodeStocks.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty) * (-1),  # negative qty
                        type="undo",
                        comment=comment,

                    )
                    success_msg += "BarcodeStocks uspesno snimljen. <br>"  # Append the success message
                except Exception as e:
                    errors.append("Problem saving to BarcodeStocks table")

        # Handle carelabel
        if carelabel != '0':

            if not errors:  # Proceed if no errors so far
                try:
                    CarelabelStocks.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty) * (-1),  # negative qty
                        type="undo",
                        comment=comment,

                    )
                    success_msg += "CarelabelStocks uspesno snimljen."  # Append the success message
                except Exception as e:
                    errors.append("Problem saving to CarelabelStocks table")

        if carelabel == '0' and barcode == '0':
            errors.append("Nije oznacen ni barcode ni carelabel")

        # If there are no errors, pass success_msg to template
        if not errors:
            return render(request, 'preparation/reduce_from_stock.html', {
                'pos': Pos.objects.filter(closed_po='Open'),
                'success_msg': success_msg
            })

        # If errors exist, pass them to the template
        return render(request, 'preparation/reduce_from_stock.html', {
            'pos': Pos.objects.filter(closed_po='Open'),
            'errors': errors
        })

    # If GET request, just render the form with open POS
    return render(request, 'preparation/reduce_from_stock.html', {
        'pos': Pos.objects.filter(closed_po='Open')
    })

def throw_away(request):
    # return HttpResponse("throw_away view is working!")

    errors = []
    success_msg = ""

    # find list of materials
    with connections['trebovanje_db'].cursor() as cursor:
        cursor.execute("SELECT DISTINCT [material] FROM [trebovanje].[dbo].[sap_coois_all] WHERE "
                       "[material] like 'A%' and wc = 'WC01'")
        materials = cursor.fetchall()
    materials = [row[0] for row in materials]
    # print(materials)

    if request.method == 'POST':
        # print(request.POST)
        material = request.POST.get('material')
        type = request.POST.get('type')
        qty = request.POST.get('qty')

        if not errors:  # Proceed if no errors so far
            try:
                ThrowAway.objects.create(

                    material=material,
                    type=type,
                    qty=int(qty)
                )
                success_msg += "ThrowAway uspesno snimljen. <br>"  # Append the success message
            except Exception as e:
                errors.append("Problem saving to ThrowAway table")

        # If there are no errors, pass success_msg to template
        if not errors:
            return render(request, 'preparation/throw_away.html', {
                'materials': materials,
                'success_msg': success_msg
            })

        # If errors exist, pass them to the template
        return render(request, 'preparation/throw_away.html', {
            'materials': materials,
            'errors': errors
        })

    # If GET request, just render the form with open POS
    return render(request, 'preparation/throw_away.html', {
        'materials': materials
    })

def leftover(request):
    # return HttpResponse("Stock view is working!")

    errors = []
    success_msg = ""

    # find list of materials
    with connections['trebovanje_db'].cursor() as cursor:
        cursor.execute(
            "SELECT DISTINCT [material] FROM [trebovanje].[dbo].[sap_coois_all] "
            "WHERE (([material] like 'A%') OR ([material] like 'ES%') OR ([material] like 'ET%') OR ([material] like 'KAFU%')) and (wc = 'WC01' OR wc = 'WC04')")
        materials = cursor.fetchall()
    materials = [row[0] for row in materials]
    # print(materials)

    # find list of skus
    with connections['posummary_db'].cursor() as cursor:
        cursor.execute(
            "SELECT DISTINCT sku FROM [posummary].[dbo].[pro] ORDER BY sku asc")
        skus = cursor.fetchall()
    skus = [row[0] for row in skus]
    # print(skus)

    if request.method == 'POST':
        print(request.POST)
        material = request.POST.get('material')
        sku = request.POST.get('sku')
        price = request.POST.get('price')
        location = request.POST.get('location')
        place = request.POST.get('place')
        qty = request.POST.get('qty')
        status = request.POST.get('status')

        if not errors:  # Proceed if no errors so far
            try:
                Leftovers.objects.create(

                    material=material,
                    sku=sku.strip(),
                    price=float(price),
                    location=location,
                    place=place,
                    qty=int(qty),
                    status=status
                )
                success_msg += "Leftover uspesno snimljen. <br>"  # Append the success message
            except Exception as e:
                errors.append("Problem saving to Leftover table")

        # If there are no errors, pass success_msg to template
        if not errors:
            return render(request, 'preparation/leftover.html', {
                'materials': materials,
                'skus': skus,
                'success_msg': success_msg
            })

        # If errors exist, pass them to the template
        return render(request, 'preparation/leftover.html', {
            'materials': materials,
            'skus': skus,
            'errors': errors
        })

    return render(request, 'preparation/leftover.html', {'materials': materials, 'skus': skus})

def transfer_to_kikinda(request):
    # return HttpResponse("transfer_to_kikinda view is working!")

    error_msg = ""
    error_msg_b = ""
    error_msg_c = ""
    success_msg = ""

    with connections['default'].cursor() as cursor:
        cursor.execute(
            """
                    SELECT  p.po_new as po
                    FROM [pos] as p
                    JOIN [172.27.161.200].[posummary].[dbo].[pro] as ps ON ps.po_new = p.po
                    WHERE p.closed_po = 'Open' AND ps.location_all = 'Kikinda'
                    ORDER BY p.created_at desc;
            """
        )
        pos_ki = cursor.fetchall()
    pos = [row[0] for row in pos_ki]
    # print(pos)

    if request.method == 'POST':
        # print(request.POST)
        po_num = request.POST.get('po')
        qty = int(request.POST.get('qty'))
        barcode = request.POST.get('barcode', '0')
        carelabel = request.POST.get('carelabel', '0')
        comment = request.POST.get('comment', '')

       # Verify if PO exists and is not closed
        try:
            po = Pos.objects.get(po=po_num)
        except ObjectDoesNotExist:
            error_msg += "Komesa doesn't exist in the PO table"
        else:
            if po.closed_po == 'Closed':
                error_msg += "PO is Closed"

        # Handle barcode
        if barcode != '0':

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(qty) as barcode_stocks
                    FROM barcode_stocks
                    WHERE ponum = %s 
                """, [po_num])
                row = cursor.fetchone()
                barcode_stocks = row[0] or 0

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(qty) as barcode_requests
                    FROM barcode_requests
                    WHERE ponum = %s AND status != 'error'
                """, [po_num])
                row = cursor.fetchone()
                barcode_requests = row[0] or 0

            if barcode_stocks - barcode_requests - qty < 0:
                error_msg_b += "Nema dovoljno barkodova na stanju <br/>"

            if error_msg_b == "":  # Proceed if no errors so far
                try:

                    # BarcodeStocks.objects.create(
                    #     po_id=po.id,
                    #     user_id=request.user.id,
                    #     ponum=po_num,
                    #     size=po.size,
                    #     qty=int(qty) * (-1),  # negative qty
                    #     type="transfer_ki",
                    #     comment=comment,
                    #
                    # )
                    BarcodeRequests.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        status='done',
                        module='kikinda',
                        type="transfer_ki",
                        comment=comment,
                    )

                    BarcodeKIStocks.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        qty_to_receive=int(qty),
                        module='PREP_SU',
                        status='to_receive',
                        type="transfer_ki",
                        comment=comment,
                    )
                    success_msg += "BarcodeKIStocks uspesno snimljen. <br>"  # Append the success message
                except Exception as e:
                    error_msg += "Problem saving to BarcodeKIStocks table"

        # Handle carelabel
        if carelabel != '0':

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(qty) as carelabel_stocks
                    FROM carelabel_stocks
                    WHERE ponum = %s 
                """, [po_num])
                row = cursor.fetchone()
                carelabel_stocks = row[0] or 0

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(qty) as carelabel_requests
                    FROM carelabel_requests
                    WHERE ponum = %s AND status != 'error'
                """, [po_num])
                row = cursor.fetchone()
                carelabel_requests = row[0] or 0

            if carelabel_stocks - carelabel_requests - qty < 0:
                error_msg_c += "Nema dovoljno carelabela na stanju <br/>"

            if error_msg_c == "":  # Proceed if no errors so far
                try:
                    # CarelabelStocks.objects.create(
                    #     po_id=po.id,
                    #     user_id=request.user.id,
                    #     ponum=po_num,
                    #     size=po.size,
                    #     qty=int(qty) * (-1),  # negative qty
                    #     type="transfer_ki",
                    #     comment=comment,
                    # )

                    CarelabelRequests.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        status='done',
                        module='kikinda',
                        type="transfer_ki",
                        comment=comment,
                    )

                    CarelabelKIStocks.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        qty_to_receive=int(qty),
                        module='PREP_SU',
                        status='to_receive',
                        type="transfer_ki",
                        comment=comment,

                    )
                    success_msg += "CarelabelKIStocks uspesno snimljen."  # Append the success message
                except Exception as e:
                    error_msg += "Problem saving to CarelabelKIStocks table"

        if carelabel == '0' and barcode == '0':
            error_msg += "Nije oznacen ni barcode ni carelabel"

        # If there are no errors, pass success_msg to template
        if not error_msg:
            return render(request, 'preparation/transfer_to_kikinda.html', {
                'pos': pos,
                'success_msg': success_msg,
                'error_msg_b': error_msg_b,
                'error_msg_c': error_msg_c,
            })

        # If errors exist, pass them to the template
        return render(request, 'preparation/transfer_to_kikinda.html', {
            'pos': pos,
            'error_msg': error_msg,
            'error_msg_b': error_msg_b,
            'error_msg_c': error_msg_c,
        })

    # If GET request, just render the form with open POS
    return render(request, 'preparation/transfer_to_kikinda.html', {
        'pos': pos
    })

def transfer_to_senta(request):
    # return HttpResponse("transfer_to_senta view is under construction!")

    error_msg = ""
    error_msg_b = ""
    error_msg_c = ""
    success_msg = ""

    with connections['default'].cursor() as cursor:
        cursor.execute(
            """
                    SELECT  p.po_new as po
                    FROM [pos] as p
                    JOIN [172.27.161.200].[posummary].[dbo].[pro] as ps ON ps.po_new = p.po
                    WHERE p.closed_po = 'Open' AND ps.location_all = 'Senta'
                    ORDER BY p.created_at desc;
            """
        )
        pos = cursor.fetchall()
    pos = [row[0] for row in pos]
    # print(pos)

    if request.method == 'POST':
        # print(request.POST)
        po_num = request.POST.get('po')
        qty = int(request.POST.get('qty'))
        barcode = request.POST.get('barcode', '0')
        carelabel = request.POST.get('carelabel', '0')
        comment = request.POST.get('comment', '')

        # Verify if PO exists and is not closed
        try:
            po = Pos.objects.get(po=po_num)
        except ObjectDoesNotExist:
            error_msg += "Komesa doesn't exist in the PO table <br/>"
        else:
            if po.closed_po == 'Closed':
                error_msg += "PO is Closed <br/>"

        # Handle barcode
        if barcode != '0':

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(qty) as barcode_stocks
                    FROM barcode_stocks
                    WHERE ponum = %s 
                """, [po_num])
                row = cursor.fetchone()
                barcode_stocks = row[0] or 0

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(qty) as barcode_requests
                    FROM barcode_requests
                    WHERE ponum = %s AND status != 'error'
                """, [po_num])
                row = cursor.fetchone()
                barcode_requests = row[0] or 0

            if barcode_stocks - barcode_requests - qty < 0:
                error_msg_b += "Nema dovoljno barkodova na stanju <br/>"

            if error_msg_b == "":  # Proceed if no errors so far
                try:

                    # BarcodeStocks.objects.create(
                    #     po_id=po.id,
                    #     user_id=request.user.id,
                    #     ponum=po_num,
                    #     size=po.size,
                    #     qty=int(qty) * (-1),  # negative qty
                    #     type="transfer_se",
                    #     comment=comment,
                    # )

                    BarcodeRequests.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        status='done',
                        module='senta',
                        type="transfer_se",
                        comment=comment,
                    )

                    BarcodeSEStocks.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        qty_to_receive=int(qty),
                        module='PREP_SU',
                        status='to_receive',
                        type="transfer_se",
                        comment=comment,

                    )
                    success_msg += "BarcodeSEStocks uspesno snimljen. <br>"  # Append the success message
                except Exception as e:
                    error_msg += "Problem saving to BarcodeSeStocks table"

        # Handle carelabel
        if carelabel != '0':

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(qty) as carelabel_stocks
                    FROM carelabel_stocks
                    WHERE ponum = %s 
                """, [po_num])
                row = cursor.fetchone()
                carelabel_stocks = row[0] or 0

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT SUM(qty) as carelabel_requests
                    FROM carelabel_requests
                    WHERE ponum = %s AND status != 'error'
                """, [po_num])
                row = cursor.fetchone()
                carelabel_requests = row[0] or 0

            if carelabel_stocks - carelabel_requests - qty < 0:
                error_msg_c += "Nema dovoljno carelabela na stanju <br/>"

            if error_msg_c == "":  # Proceed if no errors so far
                try:
                    # CarelabelStocks.objects.create(
                    #     po_id=po.id,
                    #     user_id=request.user.id,
                    #     ponum=po_num,
                    #     size=po.size,
                    #     qty=int(qty) * (-1),  # negative qty
                    #     type="transfer_se",
                    #     comment=comment,
                    # )

                    CarelabelRequests.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        status='done',
                        module='senta',
                        type="transfer_se",
                        comment=comment,
                    )
                    CarelabelSEStocks.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        qty_to_receive=int(qty),
                        module='PREP_SU',
                        status='to_receive',
                        type="transfer_se",
                        comment=comment,

                    )
                    success_msg += "CarelabelSEStocks uspesno snimljen."  # Append the success message
                except Exception as e:
                    error_msg += "Problem saving to CarelabelSeStocks table"

        # If there are no errors, pass success_msg to template
        if not error_msg:
            return render(request, 'preparation/transfer_to_senta.html', {
                'pos': pos,
                'success_msg': success_msg,
                'error_msg_b': error_msg_b,
                'error_msg_c': error_msg_c,
            })

        # If errors exist, pass them to the template
        return render(request, 'preparation/transfer_to_senta.html', {
            'pos': pos,
            'error_msg': error_msg,
            'error_msg_b': error_msg_b,
            'error_msg_c': error_msg_c,
        })

    # If GET request, just render the form with open POS
    return render(request, 'preparation/transfer_to_senta.html', {
        'pos': pos
    })

def manual_request(request):
    # return HttpResponse("manual_request view is under construction!")

    # find list of modules/lines
    with connections['bbstock_db'].cursor() as cursor:
        cursor.execute("SELECT [location] as line FROM [bbStock].[dbo].[locations]")
        lines = cursor.fetchall()
    lines = [row[0] for row in lines]
    # print(lines)

    errors = []
    success_msg = ""

    if request.method == 'POST':
        # print(request.POST)
        po_num = request.POST.get('po')
        qty = int(request.POST.get('qty') or 0)
        print(qty)
        barcode = request.POST.get('barcode', '0')
        carelabel = request.POST.get('carelabel', '0')
        modul = request.POST.get('modul')
        leader = request.POST.get('leader','')
        comment = request.POST.get('comment', '')

        # Verify if PO exists and is not closed
        try:
            po = Pos.objects.get(po=po_num)
        except ObjectDoesNotExist:
            errors.append("Komesa doesn't exist in the PO table")
        else:
            if po.closed_po == 'Closed':
                errors.append("PO is Closed")


        if qty <= 0:
            status = 'pending'
        else:
            status = 'confirmed'

        # Handle barcode
        if barcode != '0':

            if not errors:  # Proceed if no errors so far
                # try:
                    BarcodeRequests.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        module=modul,
                        leader=leader,
                        status=status,
                        type="preparation",
                        comment=comment,

                    )
                    success_msg += "Barcode manual Requests uspesno snimljen. <br>"  # Append the success message
                # except Exception as e:
                #     errors.append("Problem saving to Barcode manual Requests table")

        # Handle carelabel
        if carelabel != '0':

            if not errors:  # Proceed if no errors so far
                # try:
                    CarelabelRequests.objects.create(
                        po_id=po.id,
                        user_id=request.user.id,
                        ponum=po_num,
                        size=po.size,
                        qty=int(qty),
                        module=modul,
                        leader=leader,
                        status=status,
                        type="preparation",
                        comment=comment,

                    )
                    success_msg += "Carelabel manual Requests uspesno snimljen."  # Append the success message
                # except Exception as e:
                #     errors.append("Problem saving to Carelabel manual Requests table")

        if carelabel == '0' and barcode == '0':
            errors.append("Nije oznacen ni barcode ni carelabel")

        # If there are no errors, pass success_msg to template
        if not errors:
            return render(request, 'preparation/manual_request.html', {
                'pos': Pos.objects.filter(closed_po='Open'),
                'lines': lines,
                'success_msg': success_msg
            })

        # If errors exist, pass them to the template
        return render(request, 'preparation/manual_request.html', {
            'pos': Pos.objects.filter(closed_po='Open'),
            'lines': lines,
            'errors': errors
        })

    # If GET request, just render the form with open POS
    return render(request, 'preparation/manual_request.html', {
        'pos': Pos.objects.filter(closed_po='Open'),
        'lines': lines
    })

def leftover_table(request):

    with connections['default'].cursor() as cursor:
        cursor.execute(""" SELECT
                material, price, sku, location, place, SUM(qty) as qty
            FROM leftovers
            GROUP BY
                material, price, sku, location, place
            HAVING
            SUM(qty) != 0
        """)
        lines = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in lines]

    return render(request, 'preparation/leftover_table.html', {'data': data})

def leftover_table_all(request):

    data = Leftovers.objects.all()

    return render(request, 'preparation/leftover_table_all.html', {'data': data})

def import_file(request):
    if request.method == "POST":

        # print("POST request received")
        import_source = request.POST.get('import_source')
        uploaded_file = request.FILES.get('file')
        # print("Import source:", import_source)
        # print("Uploaded file:", uploaded_file)

        if not uploaded_file:
            messages.error(request, "No file uploaded.")
            return redirect('preparation:import_file')  # fixed name if needed

        if import_source == 'import_leftover':
            try:
                # print("Starting Leftover Import...")
                wb = openpyxl.load_workbook(uploaded_file)
                sheet = wb.active
                lines_added = 0
                lines_failed = 0

                headers = [cell.value.strip().lower() if cell.value else "" for cell in
                           next(sheet.iter_rows(min_row=1, max_row=1))]

                expected_columns = ['material', 'sku', 'price', 'location', 'place', 'qty']
                header_indices = {}
                for col in expected_columns:
                    if col in headers:
                        header_indices[col] = headers.index(col)
                    else:
                        messages.error(request, f"Missing expected column: {col}")
                        return redirect('preparation:import_file')

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        material = row[header_indices['material']]
                        sku = row[header_indices['sku']].strip()
                        price = float(row[header_indices['price']])
                        location = row[header_indices['location']]
                        place = row[header_indices['place']]
                        qty = int(row[header_indices['qty']])

                        Leftovers.objects.create(
                            material=material,
                            sku=sku,
                            price=price,
                            location=location,
                            place=place,
                            qty=qty,
                            status='ON STOCK'
                        )
                        lines_added += 1
                    except Exception as e:
                        lines_failed += 1
                        continue

                messages.success(request, f"Import finished. Added: {lines_added}, Failed: {lines_failed}")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

            # return HttpResponse("Upload received and matched import_leftover!")
            return redirect('preparation:import_file')

        elif import_source == 'import_leftover_neg':
            try:
                # print("Starting Leftover Import...")
                wb = openpyxl.load_workbook(uploaded_file)
                sheet = wb.active
                lines_added = 0
                lines_failed = 0

                headers = [cell.value.strip().lower() if cell.value else "" for cell in
                           next(sheet.iter_rows(min_row=1, max_row=1))]

                expected_columns = ['material', 'sku', 'price', 'location', 'place', 'qty']
                header_indices = {}
                for col in expected_columns:
                    if col in headers:
                        header_indices[col] = headers.index(col)
                    else:
                        messages.error(request, f"Missing expected column: {col}")
                        return redirect('preparation:import_file')

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        material = row[header_indices['material']]
                        sku = row[header_indices['sku']].strip()
                        price = float(row[header_indices['price']])
                        location = row[header_indices['location']]
                        place = row[header_indices['place']]
                        qty = int(row[header_indices['qty']]) * (-1)

                        Leftovers.objects.create(
                            material=material,
                            sku=sku,
                            price=price,
                            location=location,
                            place=place,
                            qty=qty,
                            status='USED'
                        )
                        lines_added += 1
                    except Exception as e:
                        lines_failed += 1
                        continue

                messages.success(request, f"Import finished. Added: {lines_added}, Failed: {lines_failed}")
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")

                # return HttpResponse("Upload received and matched import_leftover!")
            return redirect('preparation:import_file')

    return render(request, 'preparation/import.html')

def prep_locations(request, l_id=None):
    user = request.user
    module = user.username.lower()

    if module == "kikinda":
        location_plant_filter = "Kikinda"
    elif module == "senta":
        location_plant_filter = "Senta"
    else:
        location_plant_filter = "Subotica"

    # Default: show all if user is in "preparations" group
    if request.user.groups.filter(name="preparations").exists():
        locations = PrepLocations.objects.all()
    elif request.user.groups.filter(name="kikinda").exists():
        locations = PrepLocations.objects.filter(location_plant="Kikinda")
    elif request.user.groups.filter(name="senta").exists():
        locations = PrepLocations.objects.filter(location_plant="Senta")
    else:
        # fallback to Subotica if no specific group
        locations = PrepLocations.objects.filter(location_plant="Subotica")

    # locations = PrepLocations.objects.filter(location_plant=location_plant_filter)
    # locations = PrepLocations.objects.all()

    if request.method == "POST":
        loc_id = request.POST.get("id")
        location_name = request.POST.get("location")
        location_desc = request.POST.get("location_desc")
        location_plant = request.POST.get("location_plant")

        if loc_id and loc_id.isnumeric():
            # Edit existing
            location = get_object_or_404(PrepLocations, id=loc_id)
        else:
            # Create new
            location = PrepLocations()

        location.location = location_name
        location.location_desc = location_desc or ''
        location.location_plant = location_plant
        location.save()

        return redirect("preparation:prep_locations")

    if l_id is not None:
        if str(l_id) == "0":
            # New location form
            location = None
        else:
            location = get_object_or_404(PrepLocations, id=l_id)
        # location = get_object_or_404(PrepLocations, id=l_id)
        return render(request, 'preparation/prep_locations.html', {'location': location})


    return render(request, 'preparation/prep_locations.html', {'locations': locations})

def assign_location_to_po(request):

    if request.user.groups.filter(name="kikinda").exists():
        location_plant_filter = "Kikinda"
    elif request.user.groups.filter(name="senta").exists():
        location_plant_filter = "Senta"
    else:
        location_plant_filter = "Subotica"

    locations = PrepLocations.objects.filter(location_plant=location_plant_filter)
    # pos_list = Pos.objects.filter(location_plant=location_plant_filter) if location_plant_filter else Pos.objects.all()

    # Fetch POS data with location_plant filter if available
    with connection.cursor() as cursor:
        cursor.execute("""
                SELECT p.id, p.po_new 
                FROM [pos] as p
                JOIN [172.27.161.200].[posummary].[dbo].[pro] as ps 
                    ON ps.po_new = p.po
                WHERE p.closed_po = 'Open' AND ps.location_all = %s
                ORDER BY p.created_at desc
            """, [location_plant_filter])
        pos_rows = cursor.fetchall()

    # Convert fetched rows into a list of dictionaries
    pos_list = [{'id': row[0], 'po_new': row[1]} for row in pos_rows]

    if request.method == "POST":
        selected_pos_id = request.POST.get("pos_id")
        selected_location_id = request.POST.get("location_id")

        if not selected_pos_id or not selected_location_id:
            messages.error(request, "Both PO and Location must be selected.")
            return redirect("preparation:assign_location_to_po")

        try:
            pos = Pos.objects.get(id=selected_pos_id)
            location = PrepLocations.objects.get(id=selected_location_id)

            if location.location_plant == "Subotica":
                pos.loc_id_su = location.id
            elif location.location_plant == "Kikinda":
                pos.loc_id_ki = location.id
            elif location.location_plant == "Senta":
                pos.loc_id_se = location.id

            pos.save()
            messages.success(request, "Location assigned to PO successfully.")
        except Pos.DoesNotExist:
            messages.error(request, "Selected PO does not exist.")
        except PrepLocations.DoesNotExist:
            messages.error(request, "Selected location does not exist.")
        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")

        return redirect("preparation:assign_location_to_po")

    return render(request, "preparation/assign_location_to_po.html", {"locations": locations, "pos_list": pos_list})

def log_tables(request, action=None):
    # return HttpResponse(action)

    if action is not None:
        # return HttpResponse(action)
        if action == 'table':
            return render(request, 'preparation/log_tables.html')

        elif action == 'stock_b':
            # return HttpResponse('stock b')
            data = BarcodeStocks.objects.filter(created_at__gte=datetime.now() - timedelta(days=365)).order_by('-created_at')
            title = 'Barcode Stock'
            type = 's'
            return render(request, 'preparation/log_tables.html', {'data':data, 'title':title, 'type':type})

        elif action == 'request_b':
            # return HttpResponse('request b')
            data = BarcodeRequests.objects.filter(created_at__gte=datetime.now() - timedelta(days=365)).order_by('-created_at')
            title = 'Barcode Requests'
            type = 'r'
            return render(request, 'preparation/log_tables.html', {'data': data, 'title': title, 'type':type})

        elif action == 'stock_c':
            # return HttpResponse('stock c')
            data = CarelabelStocks.objects.filter(created_at__gte=datetime.now() - timedelta(days=365)).order_by('-created_at')
            title = 'Carelabel Stock'
            type = 's'
            return render(request, 'preparation/log_tables.html', {'data': data, 'title': title, 'type':type})

        elif action == 'request_c':
            # return HttpResponse('request c')
            data = CarelabelRequests.objects.filter(created_at__gte=datetime.now() - timedelta(days=365)).order_by('-created_at')
            title = 'Carelabel Requests'
            type = 'r'
            return render(request, 'preparation/log_tables.html', {'data': data, 'title': title, 'type':type})

        elif action == 'request_s':
            # return HttpResponse('request s')
            data = SecondQRequests.objects.filter(created_at__gte=datetime.now() - timedelta(days=365)).order_by('-created_at')
            title = 'Second quality requests'
            type = '2'
            return render(request, 'preparation/log_tables.html', {'data': data, 'title': title, 'type':type})

    else:
        # return HttpResponse(action)
        return render(request, 'preparation/log_tables.html')

#PO Table
def pos_table(request):
    pos_items = Pos.objects.filter(closed_po='Open')
    return render(request, 'preparation/pos_table.html',{'pos_items': pos_items})

def edit_pos(request, pos_id):
    pos = Pos.objects.get(id=pos_id)  # Get the specific PO entry

    if request.method == "POST":
        pos.po = request.POST.get('po')
        pos.size = request.POST.get('size')
        pos.color = request.POST.get('color')
        pos.total_order_qty = request.POST.get('total_order_qty')
        pos.comment = request.POST.get('comment')
        pos.closed_po = request.POST.get('closed_po')

        pos.save()  # Save updated PO entry
        messages.success(request, "PO updated successfully!")
        return redirect('preparation:pos_table')

    return render(request, 'preparation/edit_pos.html', {'pos': pos})

def import_pos_data(request):
    if request.method == "POST":
        try:
            with connections['posummary_db'].cursor() as cursor:
                cursor.execute(""" 
                    SELECT pro, pro_fr, material, color_desc, qty, segment, delivery_date_orig, skeda, 
                           no_lines_by_skeda, brand, status_int, season 
                    FROM pro 
                    WHERE status_int = 'Open'  
                    AND created_fr > '2023-12-31' 
                    AND deleted != 'DELETED'
                """)
                rows = cursor.fetchall()
                # print(f"Retrieved {len(rows)} rows from POS database.")

                imported_count = 0
                updated_count = 0

                with transaction.atomic():
                    for row in rows:

                        try:
                            (pro, pro_fr, material, color_desc, qty, segment, delivery_date_orig, skeda,
                             no_lines_by_skeda, brand, status_int, season) = row

                            # Ensure pro is a string
                            pro = str(pro)

                            # Parse order_code
                            order_code = pro_fr[3:]
                            po_array = order_code.split("::")
                            po = po_array[0][-7:] if len(po_array) > 0 else ""
                            po_new = po
                            size = po_array[2] if len(po_array) > 2 else ""

                            # Style and color extraction
                            style = material[:8].replace(" ", "")
                            color = material[9:13].replace(" ", "")
                            po_key = f"{po}-{size}"
                            closed = status_int
                            brand = brand if brand else "no info"

                            qty = int(qty) if qty else 0
                            delivery_date = delivery_date_orig.strftime('%Y-%m-%d')[:10] if delivery_date_orig else None

                            # ** Fetch hangtag from sap_coois_db **
                            hangtag = None
                            try:
                                with connections['trebovanje_db'].cursor() as hang_cursor:
                                    hang_cursor.execute("""
                                        SELECT material FROM trebovanje.dbo.sap_coois
                                        WHERE ((wc = 'WC01' AND material LIKE 'ES%%')
                                            OR (wc = 'WC01' AND material LIKE 'AF%%')
                                            OR (wc LIKE 'WC04%%' AND material LIKE 'ET%%'))
                                            AND po = %s
                                    """, (pro,))  # Use the safe version of pro
                                    hang_results = hang_cursor.fetchall()

                                    if hang_results:
                                        hangtag = " | ".join([row[0] for row in hang_results])
                                    # else:
                                        # print(f"No hangtag found for PO: {pro}")
                            except Exception as e:
                                # print(f"Error fetching hangtag for PO {pro_safe}: {str(e)}")
                                hangtag = "Error fetching hangtag"


                            # ** Check if PO exists in the local database **
                            # print(f"Checking if PO already exists: {pro}")
                            existing_pos = Pos.objects.filter(order_code__startswith=pro).first()

                            if existing_pos:
                                # print("Updating existing PO record...")
                                existing_pos.season = season
                                existing_pos.total_order_qty = qty
                                existing_pos.flash = segment
                                existing_pos.closed_po = closed
                                existing_pos.brand = brand
                                existing_pos.delivery_date = delivery_date
                                existing_pos.hangtag = hangtag
                                existing_pos.skeda = skeda
                                existing_pos.no_lines_by_skeda = no_lines_by_skeda
                                existing_pos.updated_at = datetime.now()
                                existing_pos.save()
                                updated_count += 1
                            else:
                                # print("Creating new PO record...")
                                pos_instance = Pos.objects.create(
                                    po_key=po_key,
                                    order_code=order_code,
                                    po=po,
                                    po_new=po_new,
                                    size=size,
                                    style=style,
                                    color=color,
                                    color_desc=color_desc,
                                    season=season,
                                    total_order_qty=qty,
                                    flash=segment,
                                    closed_po=closed,
                                    brand=brand,
                                    status=None,
                                    type=None,
                                    comment=None,
                                    delivery_date=delivery_date,
                                    hangtag=hangtag,
                                    sap_material=material,
                                    skeda=skeda,
                                    no_lines_by_skeda=no_lines_by_skeda
                                )

                                po_id = pos_instance.id
                                user_id = request.user.id if request.user.is_authenticated else None

                                BarcodeStocks.objects.create(
                                    po_id=po_id,
                                    user_id=user_id,
                                    ponum=po_new,
                                    size=size,
                                    qty=0,
                                    module=None,
                                    status=None,
                                    type='insert',
                                    comment=None,
                                    machine=None
                                )

                                # Create CarelabelStocks
                                CarelabelStocks.objects.create(
                                    po_id=po_id,
                                    user_id=user_id,
                                    ponum=po_new,
                                    size=size,
                                    qty=0,
                                    module=None,
                                    status=None,
                                    type='insert',
                                    comment=None,
                                    machine=None
                                )

                                imported_count += 1


                        except Exception as e:
                            # print(f"ERROR: {str(e)}")  # Log the error
                            messages.error(request, f"Error: {str(e)}")

        except Exception as e:
            # print(f"ERROR: {str(e)}")  # Log the error
            messages.error(request, f"Error: {str(e)}")
        messages.success(request,f"Imported {imported_count} new POS records, updated {updated_count} existing records.")
    return redirect('preparation:pos_table')

def close_pos_data(request):
    if request.method == "POST":
        try:
            closed_count = 0  # Counter for closed POs
            deleted_count = 0  # Counter for deleted POs

            # Fetch Closed PRO records from `posummary_db`
            with connections['posummary_db'].cursor() as cursor1:
                cursor1.execute("""
                    SELECT pro FROM [posummary].[dbo].[pro]
                    WHERE (status_int = 'Closed') 
                    AND created_fr > '2023-12-31'
                """)
                posummary_closed = cursor1.fetchall()

            for row in posummary_closed:
                pro_code = row[0]  # Extract `pro` value

                # Check if PO is still open in `preparation` (using standard connection)
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT id FROM pos  
                        WHERE SUBSTRING(order_code, 0, 10) = %s 
                        AND closed_po = 'Open'
                    """, [pro_code])
                    pos_closed = cursor.fetchone()  # Fetch single row

                    if not pos_closed:  # No open PO found, skip
                        continue

                    # Update `closed_po` status
                    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor.execute("""
                        UPDATE pos 
                        SET closed_po = 'Closed', updated_at = %s
                        WHERE SUBSTRING(order_code, 0, 10) = %s
                    """, [now, pro_code])

                    closed_count += 1

            # Fetch Deleted PRO records from `posummary_db`
            with connections['posummary_db'].cursor() as cursor1:
                cursor1.execute("""
                            SELECT pro FROM [posummary].[dbo].[pro]
                            WHERE (deleted = 'DELETED') 
                            AND created_fr > '2023-12-31'
                        """)
                posummary_closed = cursor1.fetchall()

            for row in posummary_closed:
                pro_code = row[0]  # Extract `pro` value

                # Check if PO is still open in `preparation` (using standard connection)
                with connection.cursor() as cursor:
                    cursor.execute("""
                                SELECT id FROM pos  
                                WHERE SUBSTRING(order_code, 0, 10) = %s 
                                AND closed_po = 'Open'
                            """, [pro_code])
                    pos_closed = cursor.fetchone()  # Fetch single row

                    if not pos_closed:  # No open PO found, skip
                        continue

                    # Update `closed_po` status
                    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    cursor.execute("""
                                UPDATE pos 
                                SET closed_po = 'Closed', updated_at = %s
                                WHERE SUBSTRING(order_code, 0, 10) = %s
                            """, [now, pro_code])

                    deleted_count += 1

            # Success message
            messages.error(request, f"Number of POs that are closed: {closed_count}, and deleted: {deleted_count}.")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    return redirect('preparation:pos_table')


