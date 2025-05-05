#core\views.py
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout, authenticate, login
import openpyxl
from core.models import *
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.contrib.auth.models import Group

app_name = 'core'

def main_page(request):

    if request.user.groups.filter(name='admins').exists():
        return redirect('core:admin_dashboard')

    elif request.user.groups.filter(name='preparations').exists():
        return redirect('preparation:preparation_dashboard')

    elif request.user.groups.filter(name='lines').exists():
        # return redirect('line:line_dashboard')
        return redirect('line:ll_login')

    elif request.user.groups.filter(name='kikinda').exists():
        return redirect('kikinda:kikinda_dashboard')

    elif request.user.groups.filter(name='senta').exists():
        return redirect('senta:senta_dashboard')

    return render(request, 'core/main_page.html')


# Login part
def login_view(request):
    if request.method == 'POST':
        # Get the form data
        username = request.POST['username']
        password = request.POST['password']

        # Authenticate the user
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Log the user in
            login(request, user)

            # Redirect user to appropriate dashboard or main page
            if user.groups.filter(name='preparations').exists():
                return redirect('preparation:preparation_dashboard')

            elif user.groups.filter(name='lines').exists():
                # return redirect('line:line_dashboard')
                return redirect('line:ll_login')

            elif request.user.groups.filter(name='kikinda').exists():
                return redirect('kikinda:kikinda_dashboard')

            elif request.user.groups.filter(name='senta').exists():
                return redirect('senta:senta_dashboard')
            else:
                return redirect('core:main_page')

        else:
            # Invalid login, show error message
            messages.error(request, 'Invalid username or password.')

    # If GET request, render the login form
    return render(request, 'core/login.html')

def logout_view(request):
    """Log the user out and redirect to login."""
    logout(request)
    return redirect('core:login')

# Admin panel
def admin_dashboard(request):

    return render(request, 'core/admin_dashboard.html')
    # is_admin = request.user.groups.filter(name='admins').exists()
    # return render(request, 'core/admin_dashboard.html', {'is_admin': is_admin})

def import_users(request):
    import_source = request.POST.get('import_source')

    if request.method == "POST" and request.FILES.get('file'):

        file = request.FILES['file']

        if import_source == 'import_users':
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                users_added = 0
                users_failed = 0

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    username, password, group_name = row

                    if CustomUser.objects.filter(username=username).exists():
                        users_failed += 1
                        continue

                    hashed_password = make_password(password)
                    user = CustomUser.objects.create(username=username, password=hashed_password)

                    group, created = Group.objects.get_or_create(name=group_name)
                    user.groups.add(group)

                    users_added += 1

                messages.success(request, f'Users imported successfully! Added: {users_added}, Failed: {users_failed}')
                return redirect('admin:core_customuser_changelist')

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                return redirect('admin:core_customuser_changelist')

        elif import_source == 'import_kistock':
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                added = 0
                failed = 0

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    commesa, accessory, stock = row

                    try:
                        pos = Pos.objects.get(po=commesa)
                    except Pos.DoesNotExist:
                        failed += 1
                        continue

                    if accessory.lower() == "barcode":
                        BarcodeKIStocks.objects.create(
                            po_id=pos.id,
                            user_id=request.user.id,
                            ponum=pos.po,
                            size=pos.size,
                            qty=stock,
                            qty_to_receive=None,
                            module='',
                            status='stock',
                            type=''
                        )
                        added += 1

                    elif accessory.lower() == "carelabel":
                        CarelabelKIStocks.objects.create(
                            po_id=pos.id,
                            user_id=request.user.id,
                            ponum=pos.po,
                            size=pos.size,
                            qty=stock,
                            qty_to_receive=None,
                            module='',
                            status='stock',
                            type=''
                        )
                        added += 1

                    else:
                        failed += 1

                messages.success(request, f'Imported stock successfully! Added: {added}, Failed: {failed}')
                return redirect('admin:core_customuser_changelist')

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                return redirect('admin:core_customuser_changelist')

        elif import_source == 'import_ki_po_loc':
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                added = 0
                failed = 0

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    commesa, location_name = row
                    try:
                        pos = Pos.objects.get(po=commesa)
                    except Pos.DoesNotExist:

                        failed += 1
                        continue
                    try:
                        location = PrepLocations.objects.get(location=location_name)
                    except PrepLocations.DoesNotExist:

                        failed += 1
                        continue

                    # Update the KI location ID
                    pos.loc_id_ki = location.id
                    pos.save()
                    added += 1
                messages.success(request, f'Location update successful! Added: {added}, Failed: {failed}')
                return redirect('admin:core_customuser_changelist')

            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
                return redirect('admin:core_customuser_changelist')

        else:
            messages.error(request, 'Invalid import source.')
            return redirect('admin:core_customuser_changelist')

    return render(request, 'core/import_users.html')
